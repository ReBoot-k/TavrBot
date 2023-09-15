import openpyxl
import re
import pickle
import difflib
import datetime
from config import FILENAME_ALL, FILENAME_CORRECTION


__version__ = "0.1.3r"

FILENAME_DATA = "data.pickle"

PATTERN_GROUP = "^[0-9][А-Я]{1,3}[0-9]+$"
PATTERN_ONE_TEACHER = r"(.*)\s([А-ЯЁ][а-яё]+(-[А-ЯЁ][а-яё]+)?\s*[А-ЯЁ]*[.]*[А-ЯЁ]*[.]*)"
PATTERN_TWO_TEACHERS = r"(.*)\s([А-ЯЁ][а-яё]+(-[А-ЯЁ][а-яё]+)?\s+[А-ЯЁ][.]*[А-ЯЁ][.]*)\s*([А-ЯЁ][а-яё]+(-[А-ЯЁ][а-яё]+)?\s+[А-ЯЁ][.]*[А-ЯЁ][.]*)"



def is_even_week(date: str) -> bool:
    """
    Проверяет, приходится ли дата на четную неделю учебного года.

    Аргументы:
        date (str): дата в формате "дд.мм"

    Возвращается:
        bool: True, если дата приходится на четную неделю, в противном случае False.
    """
    day, month = date.split(".")
    day = int(day)
    month = int(month)
    year = datetime.datetime.now().year
    date = datetime.datetime(year, month, day)
    first_weekday = datetime.datetime(year, 9, 1)
    while first_weekday.weekday() >= 5:
        first_weekday += datetime.timedelta(days=1)
    week = date.isocalendar().week
    first_week = first_weekday.isocalendar().week
    delta = week - first_week
    if delta <= 0:
        return True
    else:
        if delta % 2 == 0:
            return True
        else:
            return False


def get_closest_match(word: str, words: list, cutoff=0.6) -> str:
    """
    Дает наиболее близкое совпадение со словом из списка слов.

    Аргументы:
        word (str): подходящее слово.
        words (list): список слов для сравнения.
        cuttof (float): точность определения.

    Возвращается:
        str: Наиболее близкое совпадение со словом из списка.
    """
    lower_word = word.lower()
    lower_words = [w.lower() for w in words]
    
    matches = difflib.get_close_matches(lower_word, lower_words, cutoff=cutoff)
    if matches:
        index = lower_words.index(matches[0])
        return words[index]
    
    return None



def get_subject_teacher(string: str) -> list:
    """
    Извлекает предмет и преподавателя(ов) из строки.

    Аргументы:
        string (str): строка, содержащая предмет и преподавателя(ов).

    Возвращается:
        list: Список из двух элементов: предмета (str) и преподавателя(ов) (список str).
    """
    match = re.search(PATTERN_TWO_TEACHERS, string)
    if match:
        subject = match.group(1)
        teacher1 = match.group(2)
        teacher2 = match.group(4)
        return subject, [teacher1, teacher2]
    
    match = re.search(PATTERN_ONE_TEACHER, string)
    if match:
        subject = match.group(1)
        teacher = match.group(2)
        return subject, [teacher]

    return None


def get_schedule(wb, week_is_even) -> dict:
    """
    Получает расписание всех групп из рабочей книги.

    Аргументы:
        wb (Workbook): объект рабочей книги openpyxl.
        week_is_even (bool): четная ли неделя.

    Возвращается:
        dict: Словарь названий групп в качестве ключей и списков уроков в качестве значений.
    """
    schedule = {}
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        for row in sheet.rows:
            for cell in filter(lambda x: isinstance(x.value, str) and re.match(PATTERN_GROUP, x.value), row):
                group = cell.value

                if group not in schedule:
                    schedule[group] = []
                col_num = cell.column
                for row_num in range(cell.row+1, sheet.max_row + 1, 2):
                    merged_range = None
                    for r in sheet.merged_cells.ranges:
                        if sheet.cell(row=row_num, column=col_num).coordinate in r: 
                            merged_range = r
                            break

                    week = (not merged_range) * (not week_is_even)

                    subject_teacher = sheet.cell(row=row_num + week, column=col_num).value 
                    classroom = sheet.cell(row=row_num + week, column=col_num+1).value

                    if subject_teacher:
                        subject_teacher = get_subject_teacher(subject_teacher)
                        classroom = [classroom] if len(subject_teacher[1]) == 1 else [classroom, sheet.cell(row=row_num+1, column=col_num+1).value] 
                        if len(classroom) > 1 and not all(classroom):
                            classroom = list(filter(lambda x: True if x else False, classroom))[0].split()
                        schedule[group].append({"subject": subject_teacher[0], "teachers": subject_teacher[1], "classrooms": classroom}) 
                    else:
                        schedule[group].append({"subject": "Окно", "teachers": "", "classrooms": "диванчик"})
                    
                    schedule[group] = schedule[group][:6]

    
    return schedule


def create_schedule_by_criteria(schedule, criterion) -> dict:
    """
    Создает расписание по заданному критерию из расписания по группам.

    Аргументы:
        schedule (dict): словарь названий групп в качестве ключей и списков уроков в качестве значений.
        criterion (str): критерий, по которому группируются данные. Может быть "group", "teacher" или "subject".

    Возвращается:
        dict: Словарь значений критериев в качестве ключей и списков уроков в качестве значений.
    """
    schedule_by_criteria = {}
    for group, lessons in schedule.items():
        for pair_num, lesson in enumerate(lessons, start=1):
            if lesson:
                if criterion == "group":
                    key = group
                    value = {"subject": lesson['subject'], "teachers": lesson['teachers'], "classrooms": lesson['classrooms']}
                    value["pair_num"] = pair_num
                    if key not in schedule_by_criteria:
                        schedule_by_criteria[key] = []
                    schedule_by_criteria[key].append(value)
                elif criterion == "teacher":

                    for index, teacher in enumerate(lesson['teachers']):
                        key = teacher
                        classroom = lesson['classrooms'][index * (len(lesson['classrooms']) > 1)]
                        value = {"group": group, "subject": lesson['subject'], "classrooms": classroom}
                        value["pair_num"] = pair_num
                        if key not in schedule_by_criteria:
                            schedule_by_criteria[key] = []
                        schedule_by_criteria[key].append(value)
                elif criterion == "subject":
                    key = lesson['subject']
                    value = {"group": group, "teachers": lesson['teachers'], "classrooms": lesson['classrooms']}
                    value["pair_num"] = pair_num
                    if key not in schedule_by_criteria:
                        schedule_by_criteria[key] = []
                    schedule_by_criteria[key].append(value)
                else:
                    return None
                
    schedule_new = {}

    for key in schedule_by_criteria:
        closest_key = get_closest_match(key, list(schedule_by_criteria.keys()))
        if closest_key and difflib.SequenceMatcher(None, key, closest_key).ratio() > 0.8 and len(key) != len(closest_key):
            schedule_new[closest_key].extend(schedule_by_criteria[key])
        else:
            schedule_new[key] = schedule_by_criteria[key]

    return schedule_new




def convert_schedule_by_criteria_to_string(key, schedule, criterion) -> str:
    """
    Преобразует расписание по заданному значению критерия в строку.

    Аргументы:
        key (str): значение критерия, по которому будет отображаться расписание.
        schedule (dict): словарь значений критериев в качестве ключей и списков уроков в качестве значений.
        criterion (str): критерий, по которому будет отображаться информация. Может быть "group", "teacher" или "subject"

    Возвращается:
        str: Строковое представление расписания для заданного значения критерия.
    """
    lessons = schedule.get(key)
    if not lessons:
        return "Error"
    
    lessons.sort(key=lambda x: x["pair_num"])
    string = ""
    for pair_num, lesson in enumerate(lessons, start=1):
        if criterion == "group":
            string += f"{pair_num}. {lesson['subject']}\n"
            for sub_index, (teacher, classroom) in enumerate(zip(lesson['teachers'], lesson['classrooms']), start=1):
                string += f"{sub_index} подгруппа:\n" if len(lesson['teachers']) == 2 else ""
                string += f"Преподаватель: {teacher}\nАудитория: {classroom}\n"
        elif criterion == "teacher":
            pair_num = lesson["pair_num"]
            string += f"{pair_num}. {lesson['subject']}\n"
            string += f"Группа: {lesson['group']}\n"
            string += f"Аудитория: {lesson['classrooms']}\n"
        elif criterion == "subject":
            pair_num = lesson["pair_num"]
            string += f"{pair_num}. {lesson['group']}\n"
            for sub_index, (teacher, classroom) in enumerate(zip(lesson['teachers'], lesson['classrooms']), start=1):
                string += f"{sub_index} подгруппа:\n" if len(lesson['teachers']) == 2 else ""
                string += f"Преподаватель: {teacher}\nАудитория: {classroom}\n"
        else:
            return None
        
        string += "\n"
    
    return string


schedule = {}

def set_schedule(date=None, week_is_even=None) -> None:
    """
    Устанавливает расписание по группам из файлов с данными.

    Аргументы:
        date (bool): Дата расписания. Если не указана, то определяется по текущей дате.
        week_is_even (bool): Четность недели.

    Возвращает:
        None: Ничего не возвращает, но изменяет глобальную переменную schedule.
    """
    global schedule 

    if not date:
        date = datetime.date.today().strftime("%d.%m")
    
    try:
        with open(FILENAME_ALL, "rb") as file:
            WORKBOOK1 = openpyxl.load_workbook(file)
        with open(FILENAME_CORRECTION, "rb") as file:
            WORKBOOK2 = openpyxl.load_workbook(file)
        if week_is_even is None:
            week_is_even = is_even_week(date)
  
        schedule = get_schedule(WORKBOOK1, week_is_even)
        schedule.update(get_schedule(WORKBOOK2, week_is_even))
    except Exception:
        return


def get_result(string) -> str:
    """
    Возвращает расписание по заданной строке.

    Аргументы:
        string (str): Строка, по которой ищется расписание.

    Возвращает:
        str: Строковое представление расписания по наиболее подходящему критерию.
    """
    match = {}
    result = ""
    
    for key in ("group", "subject", "teacher"):
        match[key] = get_closest_match(string, list(create_schedule_by_criteria(schedule, key).keys()))

    for key, value in match.items():
        if value:
            result += f"Расписание для \"{value}\" ------------------:\n"
            result += convert_schedule_by_criteria_to_string(value, create_schedule_by_criteria(schedule, key), key)
            break
    
    return result


set_schedule()
