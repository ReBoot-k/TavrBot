import openpyxl
import re
import pickle
import datetime

__version__ = "0.0.10b"

with open("all.xlsx", "rb") as file:
    WORKBOOK1 = openpyxl.load_workbook(file)
with open("correction.xlsx", "rb") as file:
    WORKBOOK2 = openpyxl.load_workbook(file)

FILENAME_DATA = "data.pickle"

PATTERN_GROUP = "^[0-9][А-Я]{1,3}[0-9]+$"
PATTERN_ONE_TEACHER = r"(.*)\s([А-Я][а-я]+(-[А-Я][а-я]+)?\s+[А-Я][. ]*[А-Я][. ]*)"
PATTERN_TWO_TEACHERS = r"(.*)\s([А-Я][а-я]+(-[А-Я][а-я]+)?\s+[А-Я][. ]*[А-Я][. ]*)\s+([А-Я][а-я]+(-[А-Я][а-я]+)?\s+[А-Я][. ]*[А-Я][. ]*)"

data = {}
try:
    with open(FILENAME_DATA, "rb") as file:
        data = pickle.load(file)
except FileNotFoundError:
    pass


def get_subject_teacher(string) -> list:
    """Extracts the subject and the teacher(s) from a string.

    Args:
        string (str): A string containing the subject and the teacher(s).

    Returns:
        list: A list of two elements: the subject (str) and the teacher(s) (list of str).
    """
    match = re.search(PATTERN_TWO_TEACHERS, string)
    if match:
        subject = match.group(1)
        teacher1 = match.group(2)
        teacher2 = match.group(4)
        return subject, [teacher1, teacher2]
    
    match = re.search(PATTERN_ONE_TEACHER, string)
    if match:
        subject = match.group(1)
        teacher = match.group(2)
        return subject, [teacher]


    return None


def get_schedule(wb) -> dict:
    """Gets the schedule of all groups from a workbook.

    Args:
        wb (Workbook): An openpyxl workbook object.

    Returns:
        dict: A dictionary of group names as keys and lists of lessons as values.
    """
    schedule = {}
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        for row in sheet.rows:
            for cell in filter(lambda x: isinstance(x.value, str) and re.match(PATTERN_GROUP, x.value), row):
                group = cell.value
                schedule[group] = []
                col_num = cell.column
                for row_num in range(cell.row+1, sheet.max_row + 1, 2):
                    merged_range = None
                    for r in sheet.merged_cells.ranges:
                        if sheet.cell(row=row_num, column=col_num).coordinate in r: 
                            merged_range = r
                            break

                    week = datetime.date.today().isocalendar()[1] % 2
                    week = (not merged_range) * week 

                    subject_teacher = sheet.cell(row=row_num + week, column=col_num).value 
                    classroom = sheet.cell(row=row_num + week, column=col_num+1).value

        
                    if subject_teacher:
                        subject_teacher = get_subject_teacher(subject_teacher)
                        classroom = [classroom] if len(subject_teacher[1]) == 1 else [classroom, sheet.cell(row=row_num+1, column=col_num+1).value] # use sheet instead of ws
                        schedule[group].append({"subject": subject_teacher[0], "teachers": subject_teacher[1], "classrooms": classroom}) 
                    else:
                        schedule[group].append(None)
    
    return schedule


def convert_schedule_to_string(group, schedule) -> str:
    """Converts the schedule of a group to a string.

    Args:
        group (str): The name of the group.
        schedule (dict): A dictionary of group names as keys and lists of lessons as values.

    Returns:
        str: A string representation of the schedule of the group.
    """
    lessons = schedule.get(group)
    if not lessons:
        return "Error"
    
    while lessons and lessons[-1] is None:
        lessons.pop()

    string = ""
    for pair_num, lesson in enumerate(lessons, start=1):
        if lesson:
            string += f"{pair_num}. {lesson['subject']}\n"
            for sub_index, (teacher, classroom) in enumerate(zip(lesson['teachers'], lesson['classrooms']), start=1):
                string += f"{sub_index} подгруппа:\n" if len(lesson['teachers']) == 2 else ""
                string += f"Преподаватель: {teacher}\nАудитория: {classroom}\n"
        else: 
            string += f"{pair_num}. Окно\n"
        
        string += "\n"
    
    return string

print(f"Расписание Таврички v {__version__}\n\n")
group = input("Введите группу: ")


if group not in data:
    shedule = get_schedule(WORKBOOK1)
    shedule.update(get_schedule(WORKBOOK2))

    with open(FILENAME_DATA, "wb") as file:
            pickle.dump(shedule, file)
    data = shedule

print(convert_schedule_to_string(group, data))
